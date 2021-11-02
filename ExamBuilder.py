import openpyxl
## Below is the code which reads the excel data (Mastery_Data.xlsx) and creates a dictionary with keys student names and values the list of problems they have not mastered
wb = openpyxl.load_workbook('Mastery_Data.xlsx')
type(wb)
sheet=wb.get_sheet_by_name('Grades')
students = {}
for i in range(2,sheet.max_row+1):
     temp = []
     for j in range(2,sheet.max_column+1):
          if sheet.cell(row=i, column =j).value < 2:		# I count 2 points as mastery, you could change this to <1 if you just record mastery or no mastery
               temp.append(j-1)
     students[sheet.cell(row=i, column =1).value] = temp
## Below is the code which builds the invididual exams, store each question in the folder with the excel file, builder file
for student in students:
    f = open('exam_'+student+'.tex', 'w+')
    contents = f.readlines()
    f.close()
    contents.insert(1, '\input{Exam_number_Opener.tex}\n')
    contents.insert(2, r'\textbf{Your Name:}'+' '+ student+'\n')
    contents.insert(3, '\input{Exam_number_aftername.tex}\n')
    count=4
    for number in students[student]:
        contents.insert(count, '\input{problem'+str(number)+'.tex}\n')
        count = count + 1
    contents.insert(count + 1, '\end{document}')
    f = open('exam_'+student+'.tex', 'w')
    contents = "".join(contents)
    f.write(contents)
    f.close()
    
import os
for student in students:
    os.system('pdflatex exam_'+student+'.tex')
